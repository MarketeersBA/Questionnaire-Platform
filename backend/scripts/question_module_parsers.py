"""
Parsers for questionnaire Excel sheets → ModuleSection / ModuleQuestion structures.

Column layout (Usage & Purchase Behaveior sheets):
  0: Question code (optional, e.g. Q1)
  1: Label
  2: Arabic question text
  3: Arabic option label
  4: English option label
  5: English question text
  7: S/C or M/C type hint
  8: CATI interviewer instruction
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from backend.models import ModuleQuestion, ModuleSection, QuestionOption

# ── Stable option value slugs (analytics keys) ───────────────────────────────

USAGE_OPTION_VALUES: Dict[str, Dict[str, str]] = {
    "us_q1": {
        "Today": "today",
        "Last week": "last_week",
        "Last month": "last_month",
        "More than a month ago": "more_than_month",
    },
    "us_q2": {
        "Every day": "every_day",
        "Two or three times a week": "two_three_per_week",
        "Once a week": "once_week",
        "Every two weeks": "every_two_weeks",
        "Every three weeks": "every_three_weeks",
        "Every month": "every_month",
    },
    "us_q3": {
        "Morning": "morning",
        "Midday": "midday",
        "Night": "night",
        "Before bedtime": "before_bedtime",
        "As needed (Specify)": "as_needed",
    },
    "us_q4": {
        "Daily": "daily",
        "During outings and special occasions": "outings_occasions",
        "Before going to work/university/errands": "before_work_uni",
        "While traveling or on trips": "while_traveling",
        "When needed (Specify)": "when_needed",
    },
}

PRICING_OPTION_VALUES: Dict[str, Dict[str, str]] = {
    "cb_q1": {
        "Less than 100 EGP": "less_than_100_egp",
        "100 – 200 EGP": "100_200_egp",
        "200 – 300 EGP": "200_300_egp",
        "300 – 400 EGP": "300_400_egp",
        "More than 400 EGP": "more_than_400_egp",
    },
    "cb_q2": {
        "I buy as needed": "buy_as_needed",
        "I buy in bulk and store": "buy_bulk_store",
        "I buy based on promotions and discounts": "buy_promotions",
    },
    "cb_q3": {
        "Grocery store": "grocery_store",
        "Supermarket/Hypermarket": "supermarket_hypermarket",
        "Kiosk": "kiosk",
        "Online (Specify)": "online_other",
        "Pharmacy": "pharmacy",
        "Other (Specify)": "other",
    },
    "cb_q4": {
        "Small": "small",
        "Medium": "medium",
        "Large": "large",
        "Based on availability": "based_on_availability",
    },
}

SPECIFY_VALUES = frozenset(
    {
        "as_needed",
        "when_needed",
        "online_other",
        "other",
    }
)

PRICING_LABEL_TO_QID = {
    "monthly budget": "cb_q1",
    "stocking": "cb_q2",
    "purchasing places": "cb_q3",
    "sizes": "cb_q4",
}


def _cell(row: Tuple[Any, ...], idx: int) -> str:
    if idx >= len(row):
        return ""
    val = row[idx]
    return str(val).strip() if val is not None else ""


def _parse_type_hint(raw: str) -> str:
    normalized = raw.upper().replace(" ", "")
    if normalized in ("M/C", "MC"):
        return "mcq"
    return "scq"


def _slugify_fallback(en_label: str) -> str:
    text = en_label.lower().strip()
    text = text.replace("–", "-").replace("—", "-")
    text = re.sub(r"\(.*?\)", "", text)
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")[:64] or "option"


def _resolve_option_value(
    question_id: str,
    en_label: str,
    value_maps: Dict[str, Dict[str, str]],
) -> str:
    qmap = value_maps.get(question_id, {})
    if en_label in qmap:
        return qmap[en_label]
    return _slugify_fallback(en_label)


def _is_specify_option(value: str, en_label: str) -> bool:
    if value in SPECIFY_VALUES:
        return True
    lower = en_label.lower()
    return "(specify)" in lower or "(حدد)" in en_label


def _parse_question_block(
    *,
    question_id: str,
    label: str,
    ar_text: str,
    en_text: str,
    q_type: str,
    cati_instruction: str,
    option_rows: List[Tuple[str, str]],
    value_map: Dict[str, Dict[str, str]],
    order: int,
) -> ModuleQuestion:
    options: List[QuestionOption] = []
    for idx, (ar_label, en_label) in enumerate(option_rows):
        if not ar_label and not en_label:
            continue
        value = _resolve_option_value(question_id, en_label, value_map)
        options.append(
            QuestionOption(
                value=value,
                ar_label=ar_label,
                en_label=en_label,
                allows_specify=_is_specify_option(value, en_label),
                order=idx,
            )
        )

    return ModuleQuestion(
        question_id=question_id,
        label=label,
        type=q_type,  # type: ignore[arg-type]
        ar_text=ar_text,
        en_text=en_text,
        order=order,
        required=True,
        options=options,
        cati_instruction=cati_instruction or None,
    )


def parse_usage_sheet(ws) -> ModuleSection:
    """Parse the Usage sheet into a single section with us_q1–us_q4."""
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(
        (i for i, r in enumerate(rows) if _cell(r, 1).lower() == "label"),
        None,
    )
    if header_idx is None:
        raise ValueError("Usage sheet: header row not found")

    questions: List[ModuleQuestion] = []
    current: Optional[Dict[str, Any]] = None
    option_rows: List[Tuple[str, str]] = []
    q_order = 0

    for row in rows[header_idx + 1 :]:
        q_code = _cell(row, 0)
        label = _cell(row, 1)
        ar_text = _cell(row, 2)
        ar_opt = _cell(row, 3)
        en_opt = _cell(row, 4)
        en_text = _cell(row, 5)
        type_hint = _cell(row, 7)
        cati = _cell(row, 8)

        is_new_question = bool(q_code) or (label and ar_text and en_text)
        if is_new_question:
            if current:
                questions.append(
                    _parse_question_block(
                        question_id=current["question_id"],
                        label=current["label"],
                        ar_text=current["ar_text"],
                        en_text=current["en_text"],
                        q_type=current["q_type"],
                        cati_instruction=current["cati"],
                        option_rows=option_rows,
                        value_map=USAGE_OPTION_VALUES,
                        order=q_order,
                    )
                )
                q_order += 1

            q_num = re.sub(r"\D", "", q_code) or str(q_order + 1)
            current = {
                "question_id": f"us_q{q_num}",
                "label": label,
                "ar_text": ar_text,
                "en_text": en_text,
                "q_type": _parse_type_hint(type_hint),
                "cati": cati,
            }
            option_rows = [(ar_opt, en_opt)] if ar_opt or en_opt else []
            continue

        if ar_opt or en_opt:
            option_rows.append((ar_opt, en_opt))

    if current:
        questions.append(
            _parse_question_block(
                question_id=current["question_id"],
                label=current["label"],
                ar_text=current["ar_text"],
                en_text=current["en_text"],
                q_type=current["q_type"],
                cati_instruction=current["cati"],
                option_rows=option_rows,
                value_map=USAGE_OPTION_VALUES,
                order=q_order,
            )
        )

    return ModuleSection(
        section_id="usage",
        title_en="Usage",
        title_ar="الاستخدام",
        order=1,
        questions=questions,
    )


def parse_pricing_behavior_sheet(ws) -> ModuleSection:
    """Parse Purchase Behaveior sheet into cb_q1–cb_q4."""
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(
        (i for i, r in enumerate(rows) if _cell(r, 1).lower() == "label"),
        0,
    )

    questions: List[ModuleQuestion] = []
    current: Optional[Dict[str, Any]] = None
    option_rows: List[Tuple[str, str]] = []
    q_order = 0

    for row in rows[header_idx + 1 :]:
        label = _cell(row, 1)
        ar_text = _cell(row, 2)
        ar_opt = _cell(row, 3)
        en_opt = _cell(row, 4)
        en_text = _cell(row, 5)
        type_hint = _cell(row, 7)
        cati = _cell(row, 8)

        is_new_question = bool(label and ar_text and en_text)
        if is_new_question:
            if current:
                questions.append(
                    _parse_question_block(
                        question_id=current["question_id"],
                        label=current["label"],
                        ar_text=current["ar_text"],
                        en_text=current["en_text"],
                        q_type=current["q_type"],
                        cati_instruction=current["cati"],
                        option_rows=option_rows,
                        value_map=PRICING_OPTION_VALUES,
                        order=q_order,
                    )
                )
                q_order += 1

            qid = PRICING_LABEL_TO_QID.get(label.lower())
            if not qid:
                qid = f"cb_q{q_order + 1}"

            current = {
                "question_id": qid,
                "label": label,
                "ar_text": ar_text,
                "en_text": en_text,
                "q_type": _parse_type_hint(type_hint),
                "cati": cati,
            }
            option_rows = [(ar_opt, en_opt)] if ar_opt or en_opt else []
            continue

        if ar_opt or en_opt:
            option_rows.append((ar_opt, en_opt))

    if current:
        questions.append(
            _parse_question_block(
                question_id=current["question_id"],
                label=current["label"],
                ar_text=current["ar_text"],
                en_text=current["en_text"],
                q_type=current["q_type"],
                cati_instruction=current["cati"],
                option_rows=option_rows,
                value_map=PRICING_OPTION_VALUES,
                order=q_order,
            )
        )

    return ModuleSection(
        section_id="pricing_behavior",
        title_en="Purchase Behavior",
        title_ar="سلوك الشراء",
        order=1,
        questions=questions,
    )


def load_workbook_sheets(xlsx_path: Path) -> Tuple[Any, Any]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    usage_ws = wb["Usage"]
    pricing_ws = wb["Purchase Behaveior"]
    return usage_ws, pricing_ws
