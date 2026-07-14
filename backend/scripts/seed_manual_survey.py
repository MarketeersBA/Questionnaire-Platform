import sys
import openpyxl
from datetime import datetime
import asyncio
from bson import ObjectId

sys.stdout.reconfigure(encoding='utf-8')

# Mock backend setup for Motor
from backend.database import Database
from backend.config import settings
from backend.models import SurveyBase

async def run_migration():
    db = Database()
    db.connect()
    database = db.db

    # 1. Read the 3 files
    print("Loading BAPF_pivot...")
    wb_bapf = openpyxl.load_workbook(r'd:\Questioner\BAPF_pivot.xlsx', data_only=True)
    ws_bapf = wb_bapf['Sheet1']
    
    print("Loading deomg...")
    wb_demo = openpyxl.load_workbook(r'd:\Questioner\deomg.xlsx', data_only=True)
    ws_demo = wb_demo['Sheet1']
    
    print("Loading pivot_scalers...")
    wb_scale = openpyxl.load_workbook(r'd:\Questioner\pivot_scalers.xlsx', data_only=True)
    ws_scale = wb_scale['Sheet1']
    
    # 2. Extract Data
    bapf_data = {} # resp_id -> list of records
    all_33_brands = set()
    bapf_headers = [c.value for c in ws_bapf[1]]
    for row in ws_bapf.iter_rows(min_row=2, values_only=True):
        resp_id = row[0]
        brand = str(row[1]) if row[1] else "Unknown"
        all_33_brands.add(brand)
        if resp_id not in bapf_data: bapf_data[resp_id] = []
        bapf_data[resp_id].append(dict(zip(bapf_headers, row)))

    all_33_brands.add("Hero")
        
    demo_data = {}
    demo_headers = [c.value for c in ws_demo[1]]
    for row in ws_demo.iter_rows(min_row=2, values_only=True):
        resp_num = row[0]
        demo_data[resp_num] = dict(zip(demo_headers, row))
        
    scale_data = {}
    scale_headers = [c.value for c in ws_scale[1]]
    for row in ws_scale.iter_rows(min_row=2, values_only=True):
        resp_num = row[1]
        if resp_num not in scale_data: scale_data[resp_num] = []
        scale_data[resp_num].append(dict(zip(scale_headers, row)))
        
    # 3. Create the Survey Configuration
    survey_id = ObjectId()
    
    custom_sub_attrs = []
    numeric_columns = [
        'AfterTaste', 'Bite', 'Bitter', 'Chewing', 'Cohesive',
        'Fresh', 'HardOrSoft', 'InnerColor', 'InnerColorNatural',
        'LongLasting', 'OuterColor', 'OverallLikeness', 'OverallOuter',
        'OverallTaste', 'PurchaseIntent', 'Shape', 'Size', 'Smell',
        'SmellNatural', 'StrengthTaste', 'Sugar', 'Swallow', 'TasteNatural',
        'Texture', 'WellBalanced', 'Recommend', 'Overall', 'PurchaseIntentRealPri', 'SuitablePrice'
    ]
    for c in numeric_columns:
        custom_sub_attrs.append({"label": c, "minLabel": "Poor", "maxLabel": "Excellent"})
        
    survey_doc = {
        "_id": survey_id,
        "company_name": "Hero Taste Test (Manual Upload)",
        "template_id": "manual_seeder",
        "template_version": 1,
        "template_snapshot_schema": {},
        "template_snapshot_questions": [],
        "template_snapshot_l2": {},
        "customizations": {
            "brands": ["Hero", "Abu Auf"],
            "category": "Protein Bar",
            "modified_questions": [],
            "my_brand": "Hero"
        },
        "layer1_rules": {"extra_conditions": []},
        "google_form_id": "",
        "google_form_url": "",
        "status": "ready",
        "report_status": "idle",
        "sample_capacity": 400,
        "links_count": 400,
        "respondent_count": 400,
        "taste_test_config": {
            "research_flow": "taste_test",
            "category": "Protein Bar",
            "own_brand": "Hero",
            "language": "ar",
            "selected_attributes": [],
            "blocks": {
                 "include_awareness": True,
                 "include_usage_habits": False,
                 "include_category_behavior": False,
                 "include_brand_metrics": False,
                 "include_purchase_intent": True
            },
            "editor_settings": {
                  "allow_wording_edits": True,
                  "analyst_only": False,
            },
            "attribute_sequence": [
                {
                    "main_attribute": "sensory",
                    "source": "custom",
                    "sub_attributes": numeric_columns
                }
            ]
        },
        "blueprint": {
            "category": "Protein Bar",
            "rating_scale": 10,
            "own_brand": "Hero",
            "brands": [],
            "attributes": {},
            "custom_research_attributes": [
                {
                    "main_attribute": "sensory",
                    "sub_attributes": custom_sub_attrs
                }
            ]
        },
        "purchase_funnel": {
            "brand_list": [{"name_en": b, "name_ar": b} for b in sorted(all_33_brands)]
        },
        "type": "TasteTest",
        "created_at": datetime.utcnow()
    }
    await database.surveys.insert_one(survey_doc)
    print(f"Generated Survey ID: {survey_id}")
    
    # 4. Generate Responses
    responses_to_insert = []
    
    all_resp_ids = set(bapf_data.keys()) | set(demo_data.keys()) | set(scale_data.keys())
    
    # Iterate over all union users
    for resp_id in all_resp_ids:
        d_data = demo_data.get(resp_id, {})
        answers = {
            "name": d_data.get("Name"),
            "phone": str(d_data.get("Phone", "")),
            "gender": d_data.get("Gender"),
            "age": d_data.get("Age"),
            "__structured": {
                "flat_evaluations": [],
                "purchase_funnel": {
                    "aw_q1": [], 
                    "aw_q2": [],
                    "aw_q3": [],
                    "pb_q1": [],
                    "pb_q2": [],
                    "pb_q3": [],
                    "pb_q4": [],
                },
                "overall": {}
            }
        }
        
        flat_evals = answers["__structured"]["flat_evaluations"]
        pf_data = answers["__structured"]["purchase_funnel"]
        
        # Populate BAPF metrics
        sb_bapf = bapf_data.get(resp_id, [])
        for bp in sb_bapf:
            brand = bp.get("brand", "")
            if bp.get("TOM") == 1: pf_data["aw_q1"].append(brand)
            if bp.get("Unaided") == 1 and bp.get("TOM") != 1: pf_data["aw_q2"].append(brand)
            if bp.get("Aided") == 1: pf_data["aw_q3"].append(brand)
            
            pb1 = False
            if bp.get("ThreeMonths") == 1 or bp.get("SixMonths") == 1:
                pf_data["pb_q1"].append(brand)
                pb1 = True
                
            if bp.get("ThreeMonths") == 1: pf_data["pb_q2"].append(brand)
            if bp.get("MOU") == 1: pf_data["pb_q3"].append(brand)
            
            # Since Hero is not in BAPF, purchase intent mapped only for Abu Auf here? No, actually we inject both below.
            # But let's check Abu Auf PI in pivot_scalers instead to be unified.
            # We'll skip pb_q4 (Purchase Intent) and pb_q3 (MOU) mapping here and use the scalers for target brands.

        # Populate Sensory Scalers (inc. Hero)
        sb_scale = scale_data.get(resp_id, [])
        for sc in sb_scale:
            brand = sc.get("brand")
            if not brand: continue
            
            for key, val in sc.items():
                if key in ["None", "response_id", "brand"] or val is None:
                    continue
                    
                # Store numeric/text
                metric = key
                group = "sensory"

                if isinstance(val, str) and str(val).isdigit():
                    val = int(val)

                flat_evals.append({
                    "brand": brand,
                    "group": group,
                    "attribute": key,
                    "metric": metric,
                    "value": val,
                    "question_id": f"q_{key}"
                })
                
                # Dynamic Purchase Funnel mapping for Hero & Abu Auf
                if key == "PurchaseIntent" and isinstance(val, (int, float)):
                    if val >= 4:
                        if brand not in pf_data["pb_q4"]:
                            pf_data["pb_q4"].append(brand)
                            
            # Special Injection for "Hero" into Awareness/Usage/MOU 
            # If user evaluated Hero, they must be aware of it and tried it.
            if brand == "Hero":
                if "Hero" not in pf_data["aw_q1"] and "Hero" not in pf_data["aw_q2"] and "Hero" not in pf_data["aw_q3"]:
                    pf_data["aw_q3"].append("Hero") # Aided
                if "Hero" not in pf_data["pb_q1"]:
                    pf_data["pb_q1"].append("Hero")
                if "Hero" not in pf_data["pb_q2"]:
                    pf_data["pb_q2"].append("Hero")
                    
        # Filter pf_data duplicates
        for q, lst in pf_data.items():
            pf_data[q] = list(set(lst))

        responses_to_insert.append({
            "survey_id": str(survey_id),
            "token": f"MANUAL_{resp_id}",
            "phone": str(d_data.get("Phone", "")),
            "answers": answers,
            "source": "layer2",
            "submitted_at": datetime.utcnow()
        })
        
    await database.responses.insert_many(responses_to_insert)
    print(f"Inserted {len(responses_to_insert)} responses.")
    db.close()

if __name__ == "__main__":
    asyncio.run(run_migration())
