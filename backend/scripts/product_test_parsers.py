"""
Parse Product Test and Package Test question banks from the canonical Excel workbook.

Workbook: General_Product_Test_Evaluation.xlsx (repo root)
  - Sheet: General_Product_Test_Evaluation
  - Sheet: package test
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

PRODUCT_SHEET = "General_Product_Test_Evaluation"
PACKAGE_SHEET = "package test"


def get_diagnostic_tag(cell) -> Optional[str]:
    """Map Excel theme fill colors to diagnostic tags (PF / EM)."""
    try:
        if cell.fill and cell.fill.fgColor:
            fg = cell.fill.fgColor
            if fg.type == "theme":
                if fg.theme == 8:
                    return "PF"
                if fg.theme == 5:
                    return "EM"
    except Exception:
        pass
    return None


def _normalize_status(raw: Any) -> str:
    status_clean = str(raw or "optional").strip().lower()
    return status_clean if status_clean in ("fixed", "optional") else "optional"


def _parse_product_sheet(ws) -> List[Dict[str, Any]]:
    questions: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        cell_attr = ws.cell(row=r, column=1)
        attr = cell_attr.value
        if not attr:
            continue

        attr_type = ws.cell(row=r, column=2).value or ""
        parent = ws.cell(row=r, column=3).value or None
        q_type = ws.cell(row=r, column=4).value or "scale 1-5"
        ar_text = ws.cell(row=r, column=5).value or ""
        ar_opts = ws.cell(row=r, column=6).value or ""
        en_when = ws.cell(row=r, column=7).value or "After Use"
        status = ws.cell(row=r, column=9).value or "optional"
        en_text = ws.cell(row=r, column=10).value or ""
        en_opts = ws.cell(row=r, column=11).value or ""

        q_id = f"pt_q{len(questions) + 1:02d}"
        now = datetime.utcnow()
        questions.append({
            "question_id": q_id,
            "attribute": str(attr).strip(),
            "attribute_type": str(attr_type).strip().lower(),
            "parent_attribute": str(parent).strip() if parent else None,
            "diagnostic_tag": get_diagnostic_tag(cell_attr),
            "question_type": str(q_type).strip(),
            "ar_text": str(ar_text).strip(),
            "en_text": str(en_text).strip(),
            "ar_options": str(ar_opts).strip() if ar_opts else None,
            "en_options": str(en_opts).strip() if en_opts else None,
            "timing": str(en_when).strip(),
            "question_status": _normalize_status(status),
            "order": len(questions) + 1,
            "created_at": now,
            "updated_at": now,
        })
    return questions


def _parse_package_sheet(ws) -> List[Dict[str, Any]]:
    questions: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        cell_attr = ws.cell(row=r, column=1)
        attr = cell_attr.value
        if not attr:
            continue

        attr_type = ws.cell(row=r, column=2).value or ""
        parent = ws.cell(row=r, column=3).value or None
        q_type = ws.cell(row=r, column=4).value or "scale 1-5"
        ar_text = ws.cell(row=r, column=5).value or ""
        ar_opts = ws.cell(row=r, column=6).value or ""
        en_when = ws.cell(row=r, column=7).value or "Before Use"
        status = ws.cell(row=r, column=9).value or "optional"
        en_text = ws.cell(row=r, column=10).value or ""
        en_opts = ws.cell(row=r, column=11).value or ""

        q_id = f"pk_q{len(questions) + 1:02d}"
        now = datetime.utcnow()
        questions.append({
            "question_id": q_id,
            "attribute": str(attr).strip(),
            "attribute_type": str(attr_type).strip().lower(),
            "parent_attribute": str(parent).strip() if parent else None,
            "question_type": str(q_type).strip(),
            "ar_text": str(ar_text).strip(),
            "en_text": str(en_text).strip(),
            "ar_options": str(ar_opts).strip() if ar_opts else None,
            "en_options": str(en_opts).strip() if en_opts else None,
            "timing": str(en_when).strip(),
            "question_status": _normalize_status(status),
            "order": len(questions) + 1,
            "created_at": now,
            "updated_at": now,
        })
    return questions


def parse_product_test_workbook(xlsx_path: Path) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Load both product and package question lists from the Excel workbook."""
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"Product test Excel workbook not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    if PRODUCT_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{PRODUCT_SHEET}' not found in {xlsx_path}")
    if PACKAGE_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{PACKAGE_SHEET}' not found in {xlsx_path}")

    product_questions = _parse_product_sheet(wb[PRODUCT_SHEET])
    package_questions = _parse_package_sheet(wb[PACKAGE_SHEET])
    wb.close()
    return product_questions, package_questions
