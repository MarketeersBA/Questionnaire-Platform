"""Excel export using openpyxl (replaces C# ExcelEngine / Interop.Excel)."""

from typing import Any, List, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo


class StyleType:
    FONT = "Font"
    INTERIOR = "Interior"


class ExcelEngine:
    _instance = None

    def __new__(cls) -> "ExcelEngine":
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self) -> None:
        self._workbook: Workbook | None = None
        self._sheet = None
        self._total_sheets_no = 0
        self._range_next_start_row = 1
        self._range_start_row = 1
        self._range_row_count = 0
        self._range_col_count = 0

    def start_new_excel_app(self) -> None:
        self._workbook = Workbook()
        self._total_sheets_no = 0
        self._sheet = self._workbook.active
        self._range_next_start_row = 1

    def add_new_sheet(self) -> None:
        if self._workbook is None:
            self.start_new_excel_app()
        if self._total_sheets_no >= len(self._workbook.worksheets):
            self._sheet = self._workbook.create_sheet()
        else:
            self._sheet = self._workbook.worksheets[self._total_sheets_no]
        self._total_sheets_no += 1
        self._range_next_start_row = 1

    def bind_to_worksheet(self, row_count: int, columns_header: List[str], data_list: List[Any]) -> None:
        """Bind data from list of 1D arrays or 2D arrays to the worksheet."""
        data = self._merge_list_of_arrays(row_count, columns_header, data_list)
        self._bind_array(data)

    def bind_to_worksheet_dataframe(self, df: pd.DataFrame) -> None:
        """Bind a DataFrame to the worksheet (like C# bindToWorksheet(DataTable))."""
        rows = [[df.columns[i] for i in range(len(df.columns))]]
        for _, row in df.iterrows():
            rows.append(list(row))
        self._bind_array(rows)

    def apply_table_theme(self, name: str, show_banded_row: bool) -> None:
        # openpyxl table style names differ; use a built-in style
        if self._sheet is None:
            return
        max_row = self._range_start_row + self._range_row_count
        max_col = self._range_col_count
        ref = f"A{self._range_start_row}:{chr(64 + max_col)}{max_row}"
        try:
            tab = Table(displayName="Table1", ref=ref)
            style = TableStyleInfo(
                name=name if name in ["TableStyleMedium9", "TableStyleLight15"] else "TableStyleMedium9",
                showRowStripes=show_banded_row,
            )
            tab.tableStyleInfo = style
            self._sheet.add_table(tab)
        except Exception:
            pass
        # Auto-fit is not in openpyxl; columns stay default width

    def apply_conditional_formatting(self, value: str, style_type: str, color: Any) -> None:
        """Apply conditional formatting: cells equal to value get font/interior color."""
        if self._sheet is None:
            return
        from openpyxl.formatting.rule import CellIsRule

        for row in range(self._range_start_row + 1, self._range_start_row + self._range_row_count + 1):
            for col in range(1, self._range_col_count + 1):
                cell = self._sheet.cell(row=row, column=col)
                if str(cell.value) == value:
                    if style_type == StyleType.INTERIOR:
                        try:
                            rgb = self._long_to_rgb(int(color))
                            cell.fill = PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")
                        except Exception:
                            pass
                    elif style_type == StyleType.FONT:
                        try:
                            rgb = self._long_to_rgb(int(color))
                            cell.font = Font(color=rgb)
                        except Exception:
                            pass

    @staticmethod
    def _long_to_rgb(val: int) -> str:
        """Convert Excel long color to hex RGB."""
        b = val & 0xFF
        g = (val >> 8) & 0xFF
        r = (val >> 16) & 0xFF
        return f"{r:02X}{g:02X}{b:02X}"

    def set_number_format(self, fmt: str, start_col: int) -> None:
        if self._sheet is None:
            return
        for row in range(self._range_start_row + 1, self._range_start_row + self._range_row_count + 1):
            for col in range(start_col, self._range_col_count + 1):
                self._sheet.cell(row=row, column=col).number_format = fmt

    def _bind_array(self, data: List[List[Any]]) -> None:
        self._range_start_row = self._range_next_start_row
        self._range_row_count = len(data) - 1
        self._range_col_count = len(data[0]) if data else 0
        for r_idx, row in enumerate(data):
            for c_idx, val in enumerate(row):
                self._sheet.cell(row=self._range_start_row + r_idx, column=c_idx + 1, value=val)
        self._range_next_start_row += len(data) + 1

    def _merge_list_of_arrays(
        self, row_count: int, columns_header: List[str], data_list: List[Any]
    ) -> List[List[Any]]:
        num_cols = len(columns_header)
        result = [[None] * num_cols for _ in range(row_count + 1)]
        for i, h in enumerate(columns_header):
            result[0][i] = h
        col_idx = 0
        for item in data_list:
            if isinstance(item, list) and item and not isinstance(item[0], list):
                for k in range(row_count):
                    result[k + 1][col_idx] = item[k] if k < len(item) else None
                col_idx += 1
            elif isinstance(item, list) and item:
                first = item[0]
                if isinstance(first, list):
                    ncols = len(first)
                    for r in range(row_count):
                        for m in range(ncols):
                            if col_idx + m < num_cols and r < len(item):
                                row_data = item[r]
                                result[r + 1][col_idx + m] = row_data[m] if m < len(row_data) else None
                    col_idx += ncols
        return result

    def save(self, path: str) -> None:
        if self._workbook:
            self._workbook.save(path)

    def close(self) -> None:
        self._workbook = None
        self._sheet = None


# Singleton access like C# ExcelEngine.Instance
def get_excel_engine() -> ExcelEngine:
    return ExcelEngine()
