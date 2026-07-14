# -*- coding: utf-8 -*-
import sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'd:\Questioner\General_Product_Test_Evaluation.xlsx')

# Sheet 1
ws = wb['General_Product_Test_Evaluation']
print('=== Sheet: General_Product_Test_Evaluation ===')
for r in range(1, ws.max_row+1):
    row_data = []
    for c in range(1, ws.max_column+1):
        cell = ws.cell(row=r, column=c)
        val = cell.value if cell.value is not None else ''
        
        # Get font color
        font_color = 'default'
        try:
            if cell.font and cell.font.color:
                if cell.font.color.type == 'rgb' and cell.font.color.rgb:
                    font_color = str(cell.font.color.rgb)
                elif cell.font.color.type == 'theme':
                    font_color = f'theme:{cell.font.color.theme}'
                elif cell.font.color.type == 'indexed':
                    font_color = f'idx:{cell.font.color.indexed}'
        except:
            pass
        
        # Get fill/highlight color
        fill_color = 'none'
        try:
            if cell.fill and cell.fill.fgColor:
                if cell.fill.fgColor.type == 'rgb' and cell.fill.fgColor.rgb:
                    fill_color = str(cell.fill.fgColor.rgb)
                elif cell.fill.fgColor.type == 'theme':
                    fill_color = f'theme:{cell.fill.fgColor.theme}'
                elif cell.fill.fgColor.type == 'indexed':
                    fill_color = f'idx:{cell.fill.fgColor.indexed}'
        except:
            pass
            
        if val != '':
            row_data.append(f'[{cell.coordinate}] {val} (font:{font_color}, fill:{fill_color})')
    if row_data:
        print(' | '.join(row_data))
    else:
        print('(empty row)')
    print('---')

print()
print()

# Sheet 2
ws2 = wb['package test']
print('=== Sheet: package test ===')
for r in range(1, ws2.max_row+1):
    row_data = []
    for c in range(1, ws2.max_column+1):
        cell = ws2.cell(row=r, column=c)
        val = cell.value if cell.value is not None else ''
        
        font_color = 'default'
        try:
            if cell.font and cell.font.color:
                if cell.font.color.type == 'rgb' and cell.font.color.rgb:
                    font_color = str(cell.font.color.rgb)
                elif cell.font.color.type == 'theme':
                    font_color = f'theme:{cell.font.color.theme}'
                elif cell.font.color.type == 'indexed':
                    font_color = f'idx:{cell.font.color.indexed}'
        except:
            pass
        
        fill_color = 'none'
        try:
            if cell.fill and cell.fill.fgColor:
                if cell.fill.fgColor.type == 'rgb' and cell.fill.fgColor.rgb:
                    fill_color = str(cell.fill.fgColor.rgb)
                elif cell.fill.fgColor.type == 'theme':
                    fill_color = f'theme:{cell.fill.fgColor.theme}'
                elif cell.fill.fgColor.type == 'indexed':
                    fill_color = f'idx:{cell.fill.fgColor.indexed}'
        except:
            pass
            
        if val != '':
            row_data.append(f'[{cell.coordinate}] {val} (font:{font_color}, fill:{fill_color})')
    if row_data:
        print(' | '.join(row_data))
    else:
        print('(empty row)')
    print('---')
